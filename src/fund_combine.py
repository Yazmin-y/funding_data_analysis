import pandas as pd
from openpyxl import load_workbook
import os
import re

pd.set_option('display.max_columns', None)

rootPath = './海通证券-基金业绩排行'
excelNames = os.listdir(rootPath)  ##获得所有的文件名

for excelName in excelNames:
    # 获取工作表中所有的表格
    excel_name = rootPath + '/' + excelName
    print(excelName)

    wb = load_workbook(excel_name)
    table = wb.worksheets[0]
    date = table['K6'].value
    # print(date)
    Sheets = wb.sheetnames
    print(Sheets)

    merge_data = pd.DataFrame()
    # 建立空DataFrame，存储数据

    for sheet_index in range(len(Sheets) - 3):
        # 循环遍历所有Sheets
        print(Sheets[sheet_index+2])
        df = pd.DataFrame(pd.read_excel(excel_name, skiprows=3, header=[0, 1, 2], sheet_name=sheet_index + 2))
        df = df[df.columns.drop(list(df.filter(regex='排名')+df.filter(regex='海通')+df.filter(regex='一级分类')+df.filter(regex='二级分类')))]
        # df = df[df.columns.drop(list(df.filter(regex='海通')))]
        # df = df[df.columns.drop(list(df.filter(regex='一级分类')))]
        # df = df[df.columns.drop(list(df.filter(regex='二级分类')))]
        # df = df[df.columns.drop(list(df.filter(regex='截止')))]
        year1 = df.filter(regex='最近一年')
        year3 = df.filter(regex='最近三年')
        info1 = df.filter(regex='基本信息')
        info2 = df.filter(regex='基本情况')
        info2 = info2[info2.columns.drop(list(info2.filter(regex='截止日')))]
        clean_data = pd.concat([info1, year1, year3, info2], axis=1)
        print(clean_data.info())
        label = clean_data.filter(regex='评级分类')
        if label.empty:
            print('未找到列：评级分类')
            clean_data.insert(5, '评级分类', value='')
        neat = clean_data.filter(regex='单位净值')
        if neat.empty:
            print('未找到列：单位净值')
            clean_data.insert(6, '单位净值', value=0)
        # print(clean_data.info())
        clean_data.columns = ['基金代码', '基金简称', '最近一年收益率[%]', '最近三年收益率[%]', '成立日期', '评级分类', '单位净值（元）', '所属公司', '托管行',
                              '现任基金经理']
        print(clean_data.info())
        # print(df.info())
        merge_data = merge_data.append(clean_data)
        # pd.concat([merge_data, df])
    # print(merge_data.info())
    merge_data.loc[:, '计算截止日期'] = date
    print(merge_data.info())
    merge_data.to_excel(rootPath + '-合并/' + excelName)


rootPath = './海通证券-基金业绩排行-合并'
excelNames = os.listdir(rootPath)  ##获得所有的文件名
df_final = pd.DataFrame(
    columns=['基金代码', '基金简称', '最近一年收益率[%]', '最近三年收益率[%]', '成立日期', '评级分类', '单位净值（元）', '所属公司', '托管行', '现任基金经理', '计算截止日期'])

for excelName in excelNames:
    # 获取工作表中所有的表格
    excel_name = rootPath + '/' + excelName
    print(excelName)

    df = pd.DataFrame(pd.read_excel(excel_name))
    df['现任基金经理'] = df['现任基金经理'].str.split(',', expand=True)[0]
    print(df['现任基金经理'].head(5))
    df_final = pd.concat([df_final, df], axis=0)
df_final = df[df.columns.drop(list(df.filter(regex='Unnamed')))]
df_final.to_excel('./fund_final.xlsx', index=False)

df = pd.DataFrame(pd.read_pickle('./fund_final.pkl'))
df = df[df.columns.drop(list(df.filter(regex='Unnamed')))]
df.to_excel('./fund_final.xlsx', index=False)
print(df.head(5))
print(df.info())